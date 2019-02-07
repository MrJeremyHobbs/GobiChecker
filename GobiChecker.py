#!/usr/bin/env python3
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from tkinter import ttk
import configparser
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# local modules
import gobi
from alma import sru

# main program ################################################################
def main(*args):
    f_path = gui.openfile()
    if f_path == "":
        return
    
    # get row count
    row_count = sum(1 for line in open(f_path))

    # loop through and parse GOBI file
    gobi_file = open(f_path, 'r', encoding='utf-8')
    for line in gobi_file.readlines():
        
        # skip header
        if gui.counter == -1:
            gui.counter += 1
            continue
        
        # initiate Gobi order line object
        order = gobi.parse_line(line)
        
        # check for null lines and skipdddd
        if order.line_is_null == True:
            continue
            
        # _____________________ PERFORM SRU SEARCHES _________________________#
        
        # generate SRU urls
        iz_isbn_query = sru.make_url(zone="IZ", sru_path=config.iz_SRU_path, 
                                       query=f"alma.isbn={order.isbn}")
        iz_title_query = sru.make_url(zone="IZ", sru_path=config.iz_SRU_path, 
                                       query=f'alma.title="{order.title_clean}"')
        iz_kw_query = sru.make_url(zone="IZ", sru_path=config.iz_SRU_path, 
                                       query=f'alma.all_for_ui all "{order.kw}"')
        
        nz_isbn_query = sru.make_url(zone="NZ", sru_path=config.nz_SRU_path, 
                                       query=f"alma.isbn={order.isbn}")
        nz_title_query = sru.make_url(zone="NZ", sru_path=config.nz_SRU_path, 
                                       query=f'alma.title="{order.title_clean}"')
        nz_kw_query = sru.make_url(zone="NZ", sru_path=config.nz_SRU_path, 
                                       query=f'alma.all_for_ui all "{order.kw}"')
        
        urls = [
            iz_isbn_query,
            iz_title_query,
            iz_kw_query,
            nz_isbn_query,
            nz_title_query,
            nz_kw_query,
        ]
            
        (iz_isbn_query_resp, 
         iz_title_query_resp,
         iz_kw_query_resp,
         nz_isbn_query_resp,
         nz_title_query_resp,
         nz_kw_query_resp) = sru.searches(urls, 6)
             
        # create search objects
        iz_isbn = sru.parse(iz_isbn_query_resp, zone="IZ", 
                              inst_code=config.inst_code)
        iz_title = sru.parse(iz_title_query_resp, zone="IZ", 
                               inst_code=config.inst_code)
        iz_kw = sru.parse(iz_kw_query_resp, zone="IZ", 
                               inst_code=config.inst_code)
        nz_isbn = sru.parse(nz_isbn_query_resp, zone="NZ", 
                               inst_code=config.inst_code)
        nz_title = sru.parse(nz_title_query_resp, zone="NZ", 
                               inst_code=config.inst_code)
        nz_kw = sru.parse(nz_kw_query_resp, zone="NZ", 
                               inst_code=config.inst_code)
        
        #______________________ PARSE RESULTS ________________________________#        
        
        # IZ-ISBN search
        iz_isbn_recs_found = ""
        if iz_isbn.numberOfRecords > 0:
            iz_isbn_recs_found = "X"
        
        # IZ-Title search
        iz_title_recs_found = ""
        if iz_title.numberOfRecords > 0:
            iz_title_recs_found = "X"
        
        # IZ-KW search
        iz_kw_recs_found = ""
        if iz_kw.numberOfRecords > 0:
            iz_kw_recs_found = "X"
        
        # NZ-ISBN search
        nz_isbn_recs_found = ""
        if nz_isbn.numberOfRecords > 0:
            nz_isbn_recs_found = "X"
        
        # NZ-Title search
        nz_title_recs_found = ""
        if nz_title.numberOfRecords > 0:
            nz_title_recs_found = "X"
        
        # NZ-KW search
        nz_kw_recs_found = ""
        if nz_kw.numberOfRecords > 0:
            nz_kw_recs_found = "X"
        
        # ebook package
        have_e_holdings = ""
        if iz_isbn.have_e_holdings == True or nz_isbn.have_e_holdings == True:
            combined_e_holdings = iz_isbn.e_holdings + nz_isbn.e_holdings
            for holding in combined_e_holdings:
                combined_e_holdings = [holding.replace(" ()", "") \
                  for holding in combined_e_holdings]
            combined_e_holdings = ", ".join(combined_e_holdings)
            have_e_holdings = "X"
        
        # _____________________ GENERATE OUTPUT ______________________________#
        results = ""
        tag = ""
        
        if iz_isbn_recs_found == "" and \
           iz_title_recs_found == "" and \
           iz_kw_recs_found == "" and \
           have_e_holdings == "":
            tag = "ok_to_order"
            results = "OK to order"
            
        if iz_title_recs_found == "X":
            tag = "duplicate"
            results = "Duplicate-Title"

        if iz_kw_recs_found == "X":
            tag = "duplicate"
            results = "Duplicate-KW"
        
        if iz_isbn_recs_found == "X":
            tag = "duplicate"
            results = "Duplicate-ISBN"
            
        if have_e_holdings == "X":
            tag = "duplicate"
            results = f"Duplicate-Have Ebook ({combined_e_holdings})"
            
        # insert results into gui
        gui.counter += 1
        increment = 100 / row_count
        gui.insert_text(gui.counter, (order.isbn, order.title, order.author, 
                          order.pub_short, order.pub_year, order.binding, 
                          iz_isbn_recs_found, iz_title_recs_found, 
                          iz_kw_recs_found, nz_isbn_recs_found, 
                          nz_title_recs_found, nz_kw_recs_found, results), tag)
        gui.progress_bar.step(increment)
        continue
            
    # finish
    gui.progress_bar["value"] = 100
    gui.msgbox("Done.")
    gobi_file.close()


    
    
# Configurations ##############################################################
class configs:
    def __init__(self, configfile):
        self.configs = configs

        c_dict = configparser.ConfigParser()
        c_dict.read(configfile)
        
        self.version                 = c_dict['misc']['version']

        self.download_directory      = c_dict['misc']['download_directory'] \
                                           .replace('\\', '//')
        
        self.inst_code               = c_dict['SRU']['inst_code']
        self.iz_SRU_path             = c_dict['SRU']['iz_path']
        self.nz_SRU_path             = c_dict['SRU']['nz_path']
        
        self.log_directory           = c_dict['log']['log_directory'] \
                                           .replace('\\', '//')

# Gui #########################################################################
class gui:
    def __init__(self, master):
        self.master = master
        
        master.title("GobiChecker "+config.version)
        master.resizable(0, 0)
        master.minsize(width=1360, height=900)
        master.maxsize(width=1360, height=900)
        master.iconbitmap(".\images\logo_small.ico")

        logo = PhotoImage(file=".\images\logo_large.png")
        self.logo = Label(image=logo)
        self.logo.image = logo
        self.logo.pack()
        
        # counter
        self.counter = -1
        
        # frames
        self.top_frame = Frame(master)
        self.top_frame.pack(side='top', fill='both', expand=False)
        
        self.run_button = Button(self.top_frame, text="OPEN FILE AND RUN", 
                                                 font="Arial 14", 
                                                 command=main, 
                                                 relief="groove")
        self.run_button.pack(fill='both', side='left', expand=True)
        
        self.save_img = PhotoImage(format = 'png', file= '.\images\save_icon.png')
        self.save_button = Button(self.top_frame, text="SAVE LOG", 
                                                  image=self.save_img, 
                                                  font="Arial 14", 
                                                  command=self.save_log_xlsx, 
                                                  relief="groove")
        
        self.save_button.pack(fill='both', side='right', expand=False)
        
        self.mid_frame = Frame(master)
        self.mid_frame.pack(side='top', fill='both', expand=True)
        
        # tree view
        self.tree = ttk.Treeview(self.mid_frame)
        style = ttk.Style()
        style.theme_use('clam')
        
        # binds
        self.tree.bind('<Control-c>', self.copy_keyboard)
        self.tree.bind("<Button-3>", self.popup)
        
        # tree columns
        self.tree['columns'] = ('isbn', 'title', 'author', 'pub', 'pub_date', 
                                  'binding', 'iz_search_isbn', 
                                  'iz_search_title', 'iz_search_kw', 
                                  'nz_search_isbn', 'nz_search_title', 
                                  'nz_search_kw', 'results')
                                  
        self.tree.heading('#0', text='#', anchor='w')
        self.tree.heading('isbn', text='ISBN', anchor="w")
        self.tree.heading('title', text='Title', anchor="w")
        self.tree.heading('author', text='Author', anchor="w")
        self.tree.heading('pub', text='Publisher', anchor="w")
        self.tree.heading('pub_date', text='Date', anchor="w")
        self.tree.heading('binding', text='Binding', anchor="w")
        self.tree.heading('iz_search_isbn', text='IZ-ISBN', anchor="w")
        self.tree.heading('iz_search_title', text='IZ-Title', anchor="w")
        self.tree.heading('iz_search_kw', text='IZ-KW', anchor="w")
        self.tree.heading('nz_search_isbn', text='NZ-ISBN', anchor="w")
        self.tree.heading('nz_search_title', text='NZ-Title', anchor="w")
        self.tree.heading('nz_search_kw', text='NZ-KW', anchor="w")
        self.tree.heading('results', text='Results', anchor="w")
        
        self.tree.column("#0", width=40)
        self.tree.column("isbn", width=90)
        self.tree.column("title", width=300)
        self.tree.column("author", width=75)
        self.tree.column("pub", width=60)
        self.tree.column("pub_date", width=50)
        self.tree.column("binding", width=50)
        self.tree.column("iz_search_isbn", width=55, anchor="center")
        self.tree.column("iz_search_title", width=55, anchor="center")
        self.tree.column("iz_search_kw", width=55, anchor="center")
        self.tree.column("nz_search_isbn", width=55, anchor="center")
        self.tree.column("nz_search_title", width=55, anchor="center")
        self.tree.column("nz_search_kw", width=55, anchor="center")
        self.tree.column("results", width=363)
        
        self.tree.pack(fill="both", expand=False, side="left")
        
        # scrollbar
        v_scrollbar = ttk.Scrollbar(self.mid_frame, orient="vertical", 
                                      command=self.tree.yview)
        v_scrollbar.place(x=1346, y=26, height=376)
        self.tree.configure(yscrollcommand=v_scrollbar.set)
       
        # tags
        self.tree.tag_configure('ok_to_order', background='white')
        self.tree.tag_configure('duplicate', background='gold')
        self.tree.tag_configure('error', background='pink')
       
        # progressbar
        style.configure("red.Horizontal.TProgressbar", foreground='red', 
                          background='#2381df')
        self.progress_bar = ttk.Progressbar(master, 
                              style="red.Horizontal.TProgressbar", 
                              orient='horizontal', mode='determinate')
        self.progress_bar.pack(fill="both", expand=False, side="top")
    
        
        self.popup_menu = Menu(master, tearoff=0)
        self.popup_menu.add_command(label="Copy ISBN",
                                    command=self.copy_mouse)
        
    def popup(self, event):
        iid = self.tree.identify_row(event.y)
        if iid:
            self.tree.selection_set(iid)
            self.popup_menu.post(event.x_root, event.y_root)
        else:
            pass
        
    def copy_keyboard(self, event):
        curItem = self.tree.focus()
        item_dict = self.tree.item(curItem)
        isbn = item_dict['values'][0]
        root.clipboard_clear()
        root.clipboard_append(isbn)
        
    def copy_mouse(self):
        curItem = self.tree.focus()
        item_dict = self.tree.item(curItem)
        isbn = item_dict['values'][0]
        root.clipboard_clear()
        root.clipboard_append(isbn)
        
    def msgbox(self, msg):
        messagebox.showinfo("Attention", msg)

    def openfile(self):
        self.filename =  filedialog.askopenfilename(initialdir = config.download_directory,
                                                    title = "Select file", 
                                                    filetypes = (("TXT files",
                                                                    "*.txt"),
                                                    ("all files","*.*")))
        return self.filename
        
    def insert_text(self, counter, msg, tags):
        self.tree.insert("", "end", text=counter, values=(msg), tags=tags)
        self.tree.yview_moveto(1)
        root.update()
        
    def save_log_csv(self):
        saved_log = open(config.log_directory+"gobi_checker_log.csv", 
                                               "w", 
                                               encoding="utf-8", 
                                               newline='')
        children = self.tree.get_children()
        for child in children:
            list = self.tree.item(child)["values"]
            w = csv.writer(saved_log, quoting=csv.QUOTE_ALL)
            w.writerow(list)
        saved_log.close()
        self.msgbox("LOG SAVED SUCCESFULLY.")
        
    def save_log_xlsx(self):
        wb = Workbook()

        # grab the active worksheet
        ws = wb.active
        
        # headers
        headers = ["ISBN", "Title", "Author", "Publisher", "Date", "Binding", 
                     "IZ-ISBN", "IZ-Title", "IZ-KW", "NZ-ISBN", "NZ-Title", 
                     "NZ-KW", "Results"]
        ws.append(headers)
        
        # rows
        children = self.tree.get_children()
        for child in children:
            list = self.tree.item(child)["values"]
            list[0] = f"'{list[0]}" # add ' to isbn string
            ws.append(list)
        
        # set column widths
        ws.column_dimensions['A'].width = "20"   # isbn
        ws.column_dimensions['B'].width = "75"   # title
        ws.column_dimensions['C'].width = "40"   # author
        ws.column_dimensions['D'].width = "20"   # publisher
        ws.column_dimensions['E'].width = "15"   # date
        ws.column_dimensions['F'].width = "15"   # date
        ws.column_dimensions['G'].width = "10"   # IZ-ISBN
        ws.column_dimensions['H'].width = "10"   # IS-Title
        ws.column_dimensions['I'].width = "10"   # IZ-KW
        ws.column_dimensions['J'].width = "10"   # NZ-ISBN
        ws.column_dimensions['K'].width = "10"   # NZ-Title
        ws.column_dimensions['L'].width = "10"   # NZ-KW
        ws.column_dimensions['M'].width = "75"  # Results
        
        # freeze header
        a = ws['A2']
        ws.freeze_panes = a
        
        # set header styles
        for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
            for cell in rows:
                cell.fill = PatternFill(fgColor='FFD700', bgColor='FFFFFF', 
                                        fill_type='solid')
                cell.font = Font(size=14, 
                                 bold=True, 
                                 italic=True, 
                                 underline='single')
                cell.alignment = openpyxl.styles.Alignment(horizontal='general', 
                                                           vertical='center')

        # save the file
        wb.save(f"{config.log_directory}\gobi_checker_log.xlsx")
        self.msgbox("LOG SAVED SUCCESFULLY.")
        

# toplevel ####################################################################
config = configs('config.ini')

# gui
root = Tk()
gui = gui(root)
root.mainloop()